import re
import sys
import json
import time
import random
import gspread  # Google Sheets API
import openpyxl
import threading
from dateutil import parser
from datetime import datetime
from oauth2client.service_account import ServiceAccountCredentials
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.support import expected_conditions as EC

# Configuration Starts here

# ✅ Google Sheets API Configuration
SCOPE = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
CREDENTIALS_FILE = "creds.json"  # Ensure this file is present on the new machine

# Authenticate with Google Sheets
try:
    creds = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_FILE, SCOPE)
    client = gspread.authorize(creds)
    print("✅ Successfully authenticated with Google Sheets API.")
except Exception as e:
    print(f"❌ Error: Could not authenticate with Google Sheets API. Ensure 'credentials.json' is correct.")
    sys.exit(1)

# ✅ Load JSON Configuration
CONFIG_FILE = "config.json"

try:
    with open(CONFIG_FILE, "r", encoding="utf-8") as file:
        config = json.load(file)
    print("✅ Successfully loaded JSON configuration.")
except Exception as e:
    print(f"❌ Error: Could not read config.json - {e}")
    sys.exit(1)

#Configure Chrome options
options = Options()
#options.add_argument("--headless") # for old browser
#options.add_argument("--headless=new") # for new browser
#options.add_argument("--window-size=1920,1200")
driver = webdriver.Chrome(options=options)

# ✅ Function to format and map data
def format_row(row):
    # Extract relevant fields dynamically (handles variations in table structure)
    date_raw = row[0]  # Date (e.g., "07.02.2025 (Gen)")
    actual = row[2]  # Actual value (e.g., "143K")
    forecast = row[3]  # Forecast value (e.g., "169K")

    # Convert "143K" → "143,000"
    def convert_k_to_number(value):
        value = value.replace("K", "000") if "K" in value else value
        value = value.replace(".", ",") if "." in value else value
        return value  # If not a number, return as is

    def normalize_date(date_str):
        """
        Converts various date formats into a standard DD/MM/YYYY format.
        - Handles cases with unwanted text (e.g., "07.02.2025 (Gen)")
        - Works with European (DD.MM.YYYY), American (Feb 7, 2025), ISO (2025-02-07), and shorthand (07-02-25).
        - Ensures always returning a clean date.
        """
        try:
            # ✅ Step 1: Remove any text inside parentheses (e.g., "(Gen)", "(Mar)", etc.)
            cleaned_date = re.sub(r"\(.*?\)", "", date_str).strip()

            # ✅ Step 2: Remove extra spaces and unnecessary words
            cleaned_date = re.sub(r"[^0-9A-Za-z./\- ]", "", cleaned_date).strip()

            # ✅ Step 3: Automatically detect and parse the date
            parsed_date = parser.parse(cleaned_date, dayfirst=True)  # Always prefer European format (DD/MM/YYYY)

            # ✅ Step 4: Convert to DD/MM/YYYY format
            return parsed_date.strftime("%d/%m/%Y")

        except Exception as e:
            print(f"❌ Date parsing failed for '{date_str}': {e}")
            return date_str  # Return the original string if parsing fails

    formatted_date = normalize_date(date_raw)
    formatted_actual = convert_k_to_number(actual)
    formatted_forecast = convert_k_to_number(forecast)

    # Return values in the correct order: Date, Forecast, Actual
    return [formatted_date, formatted_forecast, formatted_actual]

def normalize_sheet_value(value):
    """
    Converts Google Sheets number format (e.g., '133.000') into a comparable plain integer (e.g., '133000').
    """
    return value.replace(".", "") if "." in value else value

MAX_RETRIES = 3  # ✅ Number of times to retry before giving up
RETRY_DELAY = 5  # ✅ Seconds to wait before retrying

def restart_webdriver():
    """ ✅ Restarts WebDriver properly """
    global driver
    print("🔄 Restarting WebDriver...")
    try:
        driver.quit()
    except:
        pass  # Ignore errors if driver is already closed
    time.sleep(3)
    driver = webdriver.Chrome(options=options)
    print("✅ WebDriver restarted successfully!")

def open_url_with_retries(driver, url):
    """
    ✅ Tries to open a URL multiple times with WebDriver restarts if needed.
    ✅ Ensures the page has loaded, the URL has changed, and JavaScript is fully rendered.
    """

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            print(f"🌍 Attempt {attempt} - Opening URL: {url}")

            # ✅ Get current URL before navigating
            old_url = driver.current_url

            # ✅ Navigate to new URL
            driver.get(url)

            # ✅ Wait until the URL changes
            WebDriverWait(driver, 10).until(EC.url_changes(old_url))

            # ✅ Wait until JavaScript fully renders the page
            WebDriverWait(driver, 10).until(
                lambda d: d.execute_script('return document.readyState') == 'complete'
            )

            # ✅ Ensure a key element is loaded (adjust as needed)
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.TAG_NAME, "table"))
            )

            print("✅ Page loaded successfully!")
            return True  # ✅ Success, exit function

        except (TimeoutException, WebDriverException) as e:
            print(f"⚠️ Error loading page (Attempt {attempt}/{MAX_RETRIES}): {e}")

            if attempt < MAX_RETRIES:
                wait_time = RETRY_DELAY + random.uniform(1, 3)
                print(f"🔄 Retrying in {wait_time:.1f} seconds...")
                time.sleep(wait_time)
            else:
                print("❌ Max retries reached. Restarting WebDriver...")
                restart_webdriver()
                return False  # ❌ Skip this URL

    return False  # ❌ If all retries fail, return False

def is_row_complete(row):
    """ ✅ Checks if all fields in a row are non-empty (i.e., data is complete). """
    return all(field.strip() != "" for field in row)

def update_config_file(config, config_file_path, spreadsheet_url, tab_name, scrape_url, new_row):
    """
    ✅ Updates the JSON config file to set the new row number after successful data fetch.
    """
    for sheet in config["spreadsheets"]:
        if sheet["url"] == spreadsheet_url:
            for tab, scraping_tasks in sheet["tabs"].items():
                if tab == tab_name:
                    for task in scraping_tasks:
                        if task["url"] == scrape_url:
                            print(f"🔄 Updating row in config: {task['row']} → {new_row}")
                            task["row"] = new_row  # Update row
                            break

    # ✅ Save the updated configuration
    try:
        with open(config_file_path, "w", encoding="utf-8") as file:
            json.dump(config, file, indent=4)
        print("✅ Config file updated successfully.")
    except Exception as e:
        print(f"❌ Failed to update config.json: {e}")

def compare_dates(new_date, sheet_date):
    """
    ✅ Compares two dates and returns:
    - 0 if they are the same
    - 1 if the new date is later (newer data)
    - -1 if the new date is earlier (should not happen)
    """
    try:
        new_dt = datetime.strptime(new_date, "%d/%m/%Y")
        sheet_dt = datetime.strptime(sheet_date, "%d/%m/%Y")

        if new_dt > sheet_dt:
            return 1  # ✅ Newer date → Move to next row
        elif new_dt == sheet_dt:
            return 0  # 🚫 Duplicate → Do not increment row
        else:
            return -1  # ❌ Something is wrong (new data should not be older)
    except Exception as e:
        print(f"⚠️ Error comparing dates: {e}")
        return -1  # Default to not increment if comparison fails

# Logic Starts Here

# ✅ Process Each Google Spreadsheet from JSON
for sheet_config in config["spreadsheets"]:
    spreadsheet_url = sheet_config["url"]

    try:
        spreadsheet = client.open_by_url(spreadsheet_url)
        print(f"\n🔹 Processing Spreadsheet: {spreadsheet.title}")
    except Exception as e:
        print(f"❌ Error: Could not access Google Spreadsheet {spreadsheet_url}. Skipping... {e}")
        continue

    # ✅ Loop Through Tabs in Each Spreadsheet
    for tab_name, scraping_tasks in sheet_config["tabs"].items():
        print(f"\n🔹 Processing Tab: {tab_name}")

        try:
            sheet = spreadsheet.worksheet(tab_name)
            print(f"  ✅ Opened tab: {tab_name}")
        except Exception as e:
            print(f"  ❌ Error: Could not find tab '{tab_name}' in Google Sheets. Skipping...")
            continue

        # ✅ Loop Through Each Scraping Task for This Tab
        for task in scraping_tasks:
            url = task["url"]
            start_row = task["row"]
            start_col = task["column"]
            fields = task["fields"]

            if not open_url_with_retries(driver, url):
                continue  # Skip this URL if it fails after retries
            print(f"  🔹 Scraping: {url} (Row: {start_row}, Column: {start_col})")

            try:
                # Click on "Cerca" button
                table = driver.find_element(By.TAG_NAME, "table")
                thead = table.find_element(By.TAG_NAME, "thead")
                headers = [th.text.strip() for th in thead.find_elements(By.TAG_NAME, "th")]
                tbody = table.find_element(By.TAG_NAME, "tbody")
                rows = tbody.find_elements(By.TAG_NAME, "tr")
                table_data = []

                # ✅ Fetch only the first two rows
                for row in rows[:2]:
                    cols = row.find_elements(By.TAG_NAME, "td")
                    row_data = [col.text.strip() for col in cols]  # Clean the text
                    if row_data:
                        table_data.append(row_data)

                # Ensure at least 2 rows were extracted
                if len(table_data) < 2 or any(len(row) == 0 for row in table_data):
                    print("❌ Not enough valid rows found in the table. Exiting.")
                    sys.exit(1)

                # ✅ Apply formatting to table_data
                table_data = list(map(format_row, table_data))

                # ✅ Get existing Google Sheet data
                existing_data = sheet.get_all_values()
                last_sheet_row = len(existing_data)  # Last row index in Google Sheets (1-based)

                # ✅ Extract the first scraped row (most recent)
                first_scraped_row = table_data[0]

                if fields == ["Date", "Actual"]:
                    first_scraped_row = [first_scraped_row[0], first_scraped_row[1]]
                    # ✅ Read the current data in the specified row to check if an update is needed
                    existing_first_row_data = existing_data[start_row-1][start_col-1:start_col] if len(existing_data) >= start_row else ["", ""]
                else:
                    # ✅ Read the current data in the specified row to check if an update is needed
                    existing_first_row_data = existing_data[start_row - 1][start_col - 1:4] if len(
                        existing_data) >= start_row else ["", "", ""]

                existing_first_row_data = [normalize_sheet_value(val) for val in existing_first_row_data]

                # ✅ Compare new date with existing date in the sheet
                date_comparison = compare_dates(first_scraped_row[0], existing_first_row_data[0])

                # ✅ Case 1: Newer date → Update `config.json` first, then write data
                if date_comparison == 1:
                    print(f"🔄 Newer data detected. Updating config.json first.")

                    # ✅ Update `config.json` row pointer **before** writing to the sheet
                    new_row = start_row + 1
                    update_config_file(config, CONFIG_FILE, spreadsheet_url, tab_name, url, new_row)
                    print(f"✅ Row updated in config.json: Next scrape will start from row {new_row}")

                    # ✅ Now write the new data to the sheet
                    for col_index, cell_value in enumerate(first_scraped_row):
                        sheet.update_cell(start_row, start_col + col_index, cell_value)
                    print(f"✅ First fetched row ({first_scraped_row[0]}) inserted at row {start_row}.")

                # ✅ Case 2: Same date → Check if values differ, update if necessary (but don't increment row)
                elif date_comparison == 0:
                    print(f"🔄 Same date detected. Checking if update is needed...")

                    if first_scraped_row != existing_first_row_data:
                        for col_index, cell_value in enumerate(first_scraped_row):
                            sheet.update_cell(start_row, start_col + col_index, cell_value)
                        print(f"✅ Data updated at row {start_row} (without moving row pointer).")
                    else:
                        print(f"🚫 No changes needed: Data is identical at row {start_row}.")

                # ✅ Case 3: Fetched date is somehow older (should not happen)
                else:
                    print(
                        f"⚠️ Warning: Fetched date is earlier than existing date. Skipping update.")

                # ✅ Extract and compare the second formatted row
                second_scraped_row = table_data[1][:3]
                prev_row_index = start_row - 1

                # ✅ Read the row before the specified row for comparison
                existing_prev_row_data = existing_data[prev_row_index - 1][start_col-1:4] if len(existing_data) >= prev_row_index else ["", "",
                                                                                                                             ""]
                existing_prev_row_data = [normalize_sheet_value(val) for val in existing_prev_row_data]
                # ✅ Overwrite the previous row **only if there are changes**
                if second_scraped_row != existing_prev_row_data:
                    for col_index, cell_value in enumerate(second_scraped_row):
                        sheet.update_cell(prev_row_index, start_col + col_index, cell_value)
                    print(f"✅ Overwritten previous row {prev_row_index} with new data ({second_scraped_row[0]}).")
                else:
                    print(f"🚫 No changes needed for previous row {prev_row_index}, data already matches.")

                ''' OPTIONAL: call -> func @ Local Excel file loading '''
            except Exception as error:
                print(error)
                print("Error! Terminating execution.")
                driver.quit()
                sys.exit(1)
    print("\nDone. Moving to the next sheet...")

driver.quit()
print("\n🔻 Scraping process completed.")

''' func @ Local Excel file loading 
    # ✅ Write data to Excel (specific region)
    # excel_filename = input("Inserisci il percorso del file: ")
    # # excel_filename = "pce_price_index.xlsx"
    # try:
    #     wb = openpyxl.load_workbook(excel_filename)  # Load existing Excel file
    #     sheet = wb.active  # Select the first sheet
    #     print(f"✅ Loaded existing Excel file: {excel_filename}")
    # except FileNotFoundError:
    #     wb = openpyxl.Workbook()  # Create a new Excel file if not found
    #     sheet = wb.active
    #     print(f"✅ Created new Excel file: {excel_filename}")
    #
    # # 🔽 **Choose where to insert data (e.g., Start from B3)**
    # start_row = int(input("Inserisci il numero della riga: "))
    # start_col = int(input("Inserisci il numero della colonna: "))
    # # start_row = 3
    # # start_col = 2  # Column B (1 = A, 2 = B, etc.)
    #
    # # ✅ Insert column headers at the specified position
    # for col_index, header in enumerate(headers):
    #     sheet.cell(row=start_row, column=start_col + col_index, value=header)
    #
    # # ✅ Insert table data at the specified position
    # for row_index, row_data in enumerate(table_data):
    #     for col_index, cell_value in enumerate(row_data):
    #         sheet.cell(row=start_row + row_index + 1, column=start_col + col_index, value=cell_value)
    #
    # # ✅ Save the updated Excel file
    # wb.save(excel_filename)
    # print(f"✅ Data successfully saved to {excel_filename} (starting at row,col -> {start_row},{start_col})") '''

''' User input
    # # ✅ Ask user for Google Sheet URL
    # spreadsheet_url = input("Inserisci l'URL del Google Sheet: ").strip()
    
    # # Validate access to the sheet
    # try:
    #     spreadsheet = client.open_by_url(spreadsheet_url)
    # 
    #     sheet_name = input("Inserisci il nome del foglio (Lascia vuoto per il primo): ").strip()
    #     if sheet_name:
    #         sheet = spreadsheet.worksheet(sheet_name)
    #     else:
    #         sheet = spreadsheet.sheet1  # Default to the first sheet
    # 
    #     print(f"✅ Successfully accessed Google Sheet: {spreadsheet.title}")
    # except Exception as e:
    #     print("❌ Error: Could not access Google Sheet. Ensure the service account email has 'Editor' permissions.")
    #     sys.exit(1)
    
    # # ✅ Ask user for row and column input
    # try:
    #     start_row = int(input("Inserisci il numero della riga: ").strip())
    #     start_col = int(input("Inserisci il numero della colonna: ").strip())
    # except ValueError:
    #     print("❌ Input non valido, devi inserire numeri...")
    #     sys.exit(1)
    
    # #Navigate to the URL
    # page = input("Inserisci l'URL per lo scraping: ")
'''
